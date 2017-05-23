# coding=utf-8
"""
"""
from __future__ import absolute_import, print_function

import hashlib
import logging
from datetime import datetime
from functools import wraps
from io import BytesIO
from operator import attrgetter
from pathlib import Path
from time import gmtime, strftime
import json
import openpyxl
import pytz
import sqlalchemy as sa
from flask import current_app, flash, g, jsonify, redirect, render_template, \
    request, session, url_for
from flask_login import current_user, login_required
from openpyxl.writer.write_only import WriteOnlyCell
from six import text_type
from werkzeug.exceptions import BadRequest, InternalServerError, NotFound
from whoosh.searching import Hit
from flask_mail import Message
from werkzeug.utils import secure_filename
import requests

from abilian.core.extensions import db
from abilian.core.models.subjects import Group, User
from abilian.core.signals import activity
from abilian.core.util import utc_dt
from abilian.i18n import _, _l
from abilian.sbe.apps.communities.security import is_manager
from abilian.sbe.apps.documents.models import Document
from abilian.services.activity import ActivityEntry
from abilian.services.security import Role
from abilian.services.auth.views import send_reset_password_instructions
from abilian.web import csrf, views
from abilian.web.action import Endpoint
from abilian.web.nav import BreadcrumbItem
from abilian.web.views import images as image_views
from abilian.core.commands.base import createuser

from .actions import register_actions
from .blueprint import Blueprint
from .forms import CommunityForm
from .models import Community, Membership
from collections import Counter
from .security import require_admin, require_manage

__all__ = ['communities']

logger = logging.getLogger(__name__)

EPOCH = datetime.fromtimestamp(0.0, tz=pytz.utc)


def seconds_since_epoch(dt):
    if not dt:
        return 0
    return int((utc_dt(dt) - EPOCH).total_seconds())


communities = Blueprint(
    "communities",
    __name__,
    set_community_id_prefix=False,
    template_folder='templates')
route = communities.route
add_url = communities.add_url_rule
communities.record_once(register_actions)


@communities.record_once
def register_context_processors(state):

    @state.app.context_processor
    def communities_context_processor():
        # helper to get an url for community image
        return dict(community_image_url=image_url)


def tab(tab_name):
    """
    Decorator for view functions to set the current "section" this view
    belongs to.
    """

    def decorator(f):

        @wraps(f)
        def set_current_tab(*args, **kwargs):
            g.current_tab = tab_name
            return f(*args, **kwargs)

        return set_current_tab

    return decorator


def default_view_kw(kw, obj, obj_type, obj_id, **kwargs):
    """
    Helper for using :func:`abilian.web.views.default_view` on objects that
    belongs to a community. This function should be used as `kw_func`::

        @default_view(blueprint, Model, kw_func=default_view_kw)
        @blueprint.route("/<object_id>")
        def view():
            ...

    """
    is_community = obj_type == Community.entity_type
    community_id = kw.get('community_id')

    if is_community or community_id is None:
        # when it's a community, default_view sets community_id to 'id', we want to
        # override with the slug value.
        if obj:
            if isinstance(obj, (Hit, dict)):
                community_id = obj.get('slug'
                                       if is_community else 'community_slug')
            elif is_community:
                community_id = obj.slug
            elif community_id is None and hasattr(obj, 'community'):
                try:
                    community_id = obj.community.slug
                except AttributeError:
                    pass

    if community_id is not None:
        kw['community_id'] = community_id
    else:
        raise ValueError('Cannot find community_id value')

    return kw


#
# Routes
#
@route("/")
@login_required
def index():
    query = Community.query
    sort_order = request.args.get('sort', u'').strip()
    if not sort_order:
        sort_order = session.get('sort_communities_order', 'alpha')

    if sort_order == 'activity':
        query = query.order_by(Community.last_active_at.desc())
    else:
        query = query.order_by(Community.name)

    session['sort_communities_order'] = sort_order

    if not current_user.has_role('admin'):
        # Filter with permissions
        query = query.join(Membership).filter(Membership.user == current_user)

    ctx = dict(my_communities=query.all(), sort_order=sort_order)
    return render_template("community/home.html", **ctx)


@route("/<string:community_id>/")
@views.default_view(
    communities, Community, 'community_id', kw_func=default_view_kw)
def community():
    return redirect(url_for("wall.index", community_id=g.community.slug))


@route("/json2")
def list_json2():
    """
    JSON endpoint, used for filling select boxes dynamically.
    """
    # TODO: make generic ?
    args = request.args

    q = args.get("q").replace("%", " ")
    if not q or len(q) < 2:
        raise BadRequest()

    query = db.session.query(Community.id, Community.name) \
        .filter(Community.name.ilike("%" + q + "%")) \
        .distinct() \
        .order_by(Community.name) \
        .limit(50)
    query_result = query.all()

    result = {'results': [{'id': r[0], 'text': r[1]} for r in query_result]}
    return jsonify(result)


# edit views
class BaseCommunityView(object):
    Model = Community
    pk = 'community_id'
    Form = CommunityForm
    base_template = 'community/_base.html'
    decorators = [require_admin]

    def init_object(self, args, kwargs):
        self.obj = g.community._model
        return args, kwargs

    def view_url(self):
        return url_for(self.view_endpoint, community_id=self.obj.slug)

    def get_form_kwargs(self):
        kwargs = super(BaseCommunityView, self).get_form_kwargs()

        image = self.obj.image
        if image and 'community' in g:
            setattr(image, 'url', image_url(self.obj, s=500))
            kwargs['image'] = image

        return kwargs


class CommunityEdit(BaseCommunityView, views.ObjectEdit):
    template = 'community/edit.html'
    title = _l("Edit community")
    decorators = views.ObjectEdit.decorators + (require_admin, tab('settings'))

    def breadcrumb(self):
        return BreadcrumbItem(label=_(u'Settings'),
                              icon='cog',
                              url=Endpoint('communities.settings',
                                           community_id=g.community.slug))

    def before_populate_obj(self):
        form = self.form
        name = form.name.data
        if name != self.obj.name:
            self.obj.rename(name)

        del form.name

        type = form.type.data
        if type != self.obj.type:
            self.obj.type = type
            self.obj.update_roles_on_folder()
        del form.type

        self.linked_group = form.linked_group.data or None
        if self.linked_group:
            self.linked_group = Group.query.get(int(self.linked_group))
        del form.linked_group

    def after_populate_obj(self):
        self.obj.group = self.linked_group


add_url(
    "/<string:community_id>/settings",
    view_func=CommunityEdit.as_view(
        'settings',
        view_endpoint='.community',
        message_success=_l(u"Community settings saved successfully.")))


class CommunityCreate(views.ObjectCreate, CommunityEdit):
    title = _l("Create community")
    decorators = views.ObjectCreate.decorators + (require_admin,)
    template = views.ObjectCreate.template
    base_template = views.ObjectCreate.base_template

    def breadcrumb(self):
        return BreadcrumbItem(label=_(u'Create new community'))

    def message_success(self):
        return _(u"Community %(name)s created successfully", name=self.obj.name)


add_url(
    '/new',
    view_func=CommunityCreate.as_view('new', view_endpoint='.community'))


class CommunityDelete(BaseCommunityView, views.ObjectDelete):
    get_form_kwargs = views.ObjectDelete.get_form_kwargs


add_url(
    "/<string:community_id>/destroy",
    methods=['POST'],
    view_func=CommunityDelete.as_view(
        'delete', message_success=_l(u"Community destroyed.")))

# Community Image
_DEFAULT_IMAGE = Path(__file__).parent / u'data' / u'community.png'
_DEFAULT_IMAGE_MD5 = hashlib.md5(_DEFAULT_IMAGE.open('rb').read()).hexdigest()
route('/_default_image')(image_views.StaticImageView.as_view(
    'community_default_image', set_expire=True, image=_DEFAULT_IMAGE))


class CommunityImageView(image_views.BlobView):
    id_arg = 'blob_id'

    def prepare_args(self, args, kwargs):
        community = g.community
        if not community:
            raise NotFound()

        kwargs[self.id_arg] = community.image.id
        # image = open(join(dirname(__file__), "data", "community.png"), 'rb')
        return super(CommunityImageView, self).prepare_args(args, kwargs)


image = CommunityImageView.as_view('image', max_size=500, set_expire=True)
route("/<string:community_id>/image")(image)


def image_url(community, **kwargs):
    """Return proper URL for image url."""
    if not community or not community.image:
        kwargs['md5'] = _DEFAULT_IMAGE_MD5
        return url_for('communities.community_default_image', **kwargs)

    kwargs['community_id'] = community.slug
    kwargs['md5'] = community.image.md5
    return url_for('communities.image', **kwargs)


def _members_query():
    """Helper used in members views."""
    last_activity_date = sa.sql.functions.max(ActivityEntry.happened_at) \
        .label('last_activity_date')
    memberships = User.query \
        .options(sa.orm.undefer('photo')) \
        .join(Membership) \
        .outerjoin(ActivityEntry,
                   sa.sql.and_(ActivityEntry.actor_id == User.id,
                               ActivityEntry.target_id == Membership.community_id)) \
        .filter(Membership.community == g.community, User.can_login == True) \
        .add_columns(Membership.id,
                     Membership.role,
                     last_activity_date) \
        .group_by(User, Membership.id, Membership.role) \
        .order_by(User.last_name.asc(), User.first_name.asc())

    return memberships


def _wizard_check_query(emails,is_csv=False):

    if is_csv:
        csv_data = emails
        existing_account_csv_roles = {user["email"]:user["role"] for user in csv_data}
        emails = [user["email"] for user in emails]

    emails = [email.strip() for email in emails]

    already_member_emails = [member.email for member in g.community.members if member.email in emails]
    not_member_emails = set(emails) - set(already_member_emails)

    existing_members_objects = filter(lambda user: user.email in already_member_emails, g.community.members)

    existing_accounts_objects = User.query.filter(User.email.in_(not_member_emails)).all()
    existing_account_emails = [user.email for user in existing_accounts_objects]

    emails_without_account = set(not_member_emails) - set(existing_account_emails)

    accounts_list = []
    for user in existing_accounts_objects:
            account = {}
            account["email"] = user.email
            account["first_name"] = user.first_name
            account["last_name"] = user.last_name
            account["role"] = existing_account_csv_roles[user.email] if is_csv else "member"
            account["status"] = "existing"
            accounts_list.append(account)

    if is_csv:
        emails_without_account = [csv_account for csv_account in csv_data if csv_account["email"] in emails_without_account]
        existing_accounts_objects = {"account_objects":existing_accounts_objects,"csv_roles":existing_account_csv_roles}

        for csv_account in emails_without_account:
                account = {}
                account["email"] = csv_account["email"]
                account["first_name"] = csv_account["first_name"]
                account["last_name"] = csv_account["last_name"]
                account["role"] = csv_account["role"]
                account["status"] = "new"
                accounts_list.append(account)
    else:
        for email in emails_without_account:
                account = {}
                account["email"] = email
                account["first_name"] = ""
                account["last_name"] = ""
                account["role"] = "member"
                account["status"] = "new"
                accounts_list.append(account)

    return existing_accounts_objects, existing_members_objects, accounts_list


@route("/<string:community_id>/members")
@tab('members')
def members():
    g.breadcrumb.append(BreadcrumbItem(
        label=_(u'Members'),
        url=Endpoint('communities.members', community_id=g.community.slug))
    )
    memberships = _members_query().all()
    community_threads_users = [thread.creator for thread in g.community.threads]
    threads_count = Counter(community_threads_users)

    return render_template(
        "community/members.html",
        seconds_since_epoch=seconds_since_epoch,
        is_manager=is_manager(user=current_user),
        memberships=memberships,
        threads_count=threads_count,
        csrf_token=csrf.field())


@route("/<string:community_id>/members/wizard/step1")
@tab('members')
def add_member_emails_wizard():
    g.breadcrumb.append(BreadcrumbItem(
        label=_(u'Members'),
        url=Endpoint('communities.members', community_id=g.community.slug))
    )

    return render_template(
        "community/wizard_add_emails.html",
        seconds_since_epoch=seconds_since_epoch,
        csrf_token=csrf.field())


def wizard_read_csv(csv):
    if request.method == 'POST':
        contents = csv.readlines()
        new_accounts = []
        for line in contents:
            account = {}
            data = line.split(";")
            account["email"] = data[0].strip()
            account["first_name"] = data[1].strip()
            account["last_name"] = data[2].strip()
            account["role"] = data[3].strip()
            new_accounts.append(account)
        return new_accounts


@route("/<string:community_id>/members/wizard/step2", methods=['GET','POST'])
@csrf.protect
@tab('members')
def check_members_wizard():
    if request.method == "POST":
        g.breadcrumb.append(BreadcrumbItem(
            label=_(u'Members'),
            url=Endpoint('communities.members', community_id=g.community.slug))
        )

        is_csv = False
        if request.form.get("wizard-emails"):
            wizard_emails = request.form.get("wizard-emails").split(",")
            existing_accounts_object,existing_members_objects,final_email_list = _wizard_check_query(wizard_emails)
            final_email_list_json = json.dumps(final_email_list)
        else:
            is_csv = True
            accounts_data = wizard_read_csv(request.files['csv_file'])
            existing_accounts,existing_members_objects,final_email_list = _wizard_check_query(accounts_data,is_csv=True)
            existing_accounts_object = existing_accounts["account_objects"]
            existing_accounts_csv_roles = existing_accounts["csv_roles"]
            final_email_list_json = json.dumps(final_email_list)

        if not final_email_list:
            flash(_(u"No new members were found"), 'warning')
            return redirect(url_for(".add_member_emails_wizard", community_id=g.community.slug))

        return render_template(
            "community/wizard_check_members.html",
            existing_accounts_object=existing_accounts_object,
            csv_roles=existing_accounts_csv_roles if is_csv else False,
            wizard_emails=final_email_list_json,
            existing_members_objects=existing_members_objects,
            csrf_token=csrf.field())

    return redirect(url_for(".members", community_id=community.slug))


@route("/<string:community_id>/members/wizard/step3", methods=['GET','POST'])
@csrf.protect
@tab('members')
def new_accounts_wizard():
    if request.method == "POST":
        g.breadcrumb.append(BreadcrumbItem(
            label=_(u'Members'),
            url=Endpoint('communities.members', community_id=g.community.slug))
        )

        wizard_emails = request.form.get("wizard-emails")
        wizard_accounts = json.loads(wizard_emails)

        wizard_existing_account = {}
        new_accounts = []

        for user in wizard_accounts:
            if user["status"] == "existing":
                wizard_existing_account[user["email"]] = user["role"]

            elif user["status"] == "new":
                new_accounts.append(user)

        existing_account = json.dumps(wizard_existing_account)

        return render_template(
            "community/wizard_new_accounts.html",
            existing_account=existing_account,
            new_accounts=new_accounts,
            csrf_token=csrf.field())

    return redirect(url_for(".members", community_id=community.slug))


@route("/<string:community_id>/members/wizard/complete", methods=['POST'])
@csrf.protect
def wizard_saving():
    community = g.community._model
    existing_accounts = request.form.get("existing_account")
    existing_accounts = json.loads(existing_accounts)
    new_accounts = request.form.get("new_accounts")
    new_accounts = json.loads(new_accounts)

    if not (existing_accounts or new_accounts):
        flash(_(u"No new members were found"), 'warning')
        return redirect(url_for(".members", community_id=g.community.slug))

    if existing_accounts:
        for email,role in existing_accounts.iteritems():
            user = User.query.filter(User.email == email).first()
            community.set_membership(user, role)

        app = current_app._get_current_object()
        activity.send(app, actor=user, verb="join", object=community)

        db.session.commit()

    if new_accounts:
        for account in new_accounts:
            email = account["email"]
            first_name = account["first_name"]
            last_name = account["last_name"]
            role = account["role"]

            user = User(
                email=email,
                last_name=last_name,
                first_name=first_name,
                can_login=True)
            db.session.add(user)

            community.set_membership(user, role)
            app = current_app._get_current_object()
            activity.send(app, actor=user, verb="join", object=community)
            db.session.commit()

            send_reset_password_instructions(user)

    flash(_(u"Members added Successfully"), 'success')
    return redirect(url_for(".members", community_id=community.slug))


@route("/<string:community_id>/members", methods=["POST"])
@csrf.protect
@require_manage
def members_post():
    community = g.community._model
    action = request.form.get("action")

    user_id = request.form.get("user")
    if not user_id:
        flash(_(u"You must provide a user."), 'error')
        return redirect(url_for(".members", community_id=community.slug))
    user_id = int(user_id)
    user = User.query.get(user_id)

    if action in ('add-user-role', 'set-user-role'):
        role = request.form.get("role").lower()

        community.set_membership(user, role)

        if action == 'add-user-role':
            app = current_app._get_current_object()
            activity.send(app, actor=user, verb="join", object=community)

        db.session.commit()
        return redirect(url_for('.members', community_id=community.slug))

    elif action == 'delete':
        membership_id = int(request.form['membership'])
        membership = Membership.query.get(membership_id)
        if membership.user_id != user_id:
            raise InternalServerError()

        community.remove_membership(user)

        app = current_app._get_current_object()
        activity.send(app, actor=user, verb="leave", object=community)

        db.session.commit()
        return redirect(url_for(".members", community_id=community.slug))

    else:
        raise BadRequest('Unknown action: {}'.format(repr(action)))


MEMBERS_EXPORT_HEADERS = [
    _l(u'Name'),
    _l(u'email'),
    _l(u'Last activity in this community'),
    _l(u'Role'),
]

MEMBERS_EXPORT_ATTRS = ['User', 'User.email', 'last_activity_date', 'role']

HEADER_FONT = openpyxl.styles.Font(bold=True)
HEADER_ALIGN = openpyxl.styles.Alignment(
    horizontal='center', vertical='top', wrapText=True)
XLSX_MIME = u'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


@route("/<string:community_id>/members/excel")
@tab('members')
def members_excel_export():
    community = g.community
    attributes = [attrgetter(a) for a in MEMBERS_EXPORT_ATTRS]
    BaseModel = current_app.db.Model
    wb = openpyxl.Workbook()

    if wb.worksheets:
        wb.remove_sheet(wb.active)

    ws_title = _(u'%(community)s members', community=community.name)
    ws_title = ws_title.strip()
    if len(ws_title) > 31:
        # sheet title cannot exceed 31 char. max length
        ws_title = ws_title[:30] + u'…'
    ws = wb.create_sheet(title=ws_title)
    row = 0
    cells = []

    cols_width = []
    for col, label in enumerate(MEMBERS_EXPORT_HEADERS, 1):
        value = text_type(label)
        cell = WriteOnlyCell(ws, value=value)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cells.append(cell)
        cols_width.append(len(value) + 1)

    ws.append(cells)

    for membership_info in _members_query().all():
        row += 1
        cells = []
        for col, getter in enumerate(attributes):
            value = None
            try:
                value = getter(membership_info)
            except AttributeError:
                pass

            if isinstance(value, (BaseModel, Role)):
                value = text_type(value)

            cell = WriteOnlyCell(ws, value=value)
            cells.append(value)

            # estimate width
            value = text_type(cell.value)
            width = max(len(l) for l in value.split(u'\n')) + 1
            cols_width[col] = max(width, cols_width[col])

        ws.append(cells)

    # adjust columns width
    MIN_WIDTH = 3
    MAX_WIDTH = openpyxl.utils.units.BASE_COL_WIDTH * 4

    for idx, width in enumerate(cols_width, 1):
        letter = openpyxl.utils.get_column_letter(idx)
        width = min(max(width, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[letter].width = width

    fd = BytesIO()
    wb.save(fd)
    fd.seek(0)

    response = current_app.response_class(fd, mimetype=XLSX_MIME)

    filename = u'{}-members-{}.xlsx'.format(
        community.slug, strftime("%d:%m:%Y-%H:%M:%S", gmtime()))
    response.headers['content-disposition'] = \
        u'attachment;filename="{}"'.format(filename)

    return response


#
# Hack to redirect from urls used by the search engine.
#
@route("/doc/<int:doc_id>")
def doc(doc_id):
    doc = Document.query.get(doc_id)

    if doc is None:
        raise NotFound()

    folder = doc.parent
    while True:
        parent = folder.parent
        if parent.is_root_folder:
            break
        folder = parent
    target_community = Community.query \
        .filter(Community.folder_id == folder.id) \
        .one()
    location = url_for(
        "documents.document_view",
        community_id=target_community.slug,
        doc_id=doc.id)
    return redirect(location)
